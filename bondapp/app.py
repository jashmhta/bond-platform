import io
import json
import os
import sys
import glob

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
if BASE_DIR not in sys.path:
    sys.path.insert(0, BASE_DIR)

from datetime import date, datetime
from functools import wraps

from flask import (
    Flask, jsonify, render_template, request, redirect, url_for,
    session, send_file, flash, abort,
)
import openpyxl

import db
import models
import quotesheet

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY") or "bondlab-demo-secret-2026"


class _StripVercelPrefix:
    def __init__(self, wsgi_app):
        self.wsgi_app = wsgi_app

    def __call__(self, environ, start_response):
        pi = environ.get("PATH_INFO", "")
        if pi.startswith("/api/index"):
            environ["PATH_INFO"] = pi[len("/api/index"):] or "/"
        return self.wsgi_app(environ, start_response)


app.wsgi_app = _StripVercelPrefix(app.wsgi_app)
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "admin123")
BASE_URL = os.environ.get("BASE_URL", "").rstrip("/")

db.init_db()


def login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get("admin"):
            return redirect(url_for("admin_login", next=request.path))
        return f(*args, **kwargs)
    return wrapper


def bond_public(b):
    b = dict(b)
    b["tiny_link"] = f"/b/{b['slug']}"
    return b


def make_slug(security_name, exclude_id=None):
    base = models.slugify(security_name)
    slug = base
    i = 2
    while True:
        row = db.get_bond(slug=slug)
        if row is None or (exclude_id and row["id"] == exclude_id):
            return slug
        slug = f"{base}-{i}"
        i += 1


# ---------------------------------------------------------------- public
@app.route("/")
def index():
    q = request.args.get("q", "").strip()
    category = request.args.get("category", "").strip()
    btype = request.args.get("type", "").strip()
    rating = request.args.get("rating", "").strip()
    sort = request.args.get("sort", "yield")
    bonds = db.list_bonds(q=q or None, category=category or None, btype=btype or None, rating=rating or None, sort=sort)
    all_b = db.get_all_bonds()
    categories = sorted({b["issuer_category"] for b in all_b if b["issuer_category"]})
    types = sorted({b["type"] for b in all_b if b["type"]})
    ratings = sorted({b["credit_rating"] for b in all_b if b["credit_rating"]})
    stats = {
        "count": len(all_b),
        "min_yield": min((b["offer_yield"] for b in all_b if b["offer_yield"]), default=0),
        "max_yield": max((b["offer_yield"] for b in all_b if b["offer_yield"]), default=0),
    }
    return render_template(
        "index.html", bonds=[bond_public(b) for b in bonds], q=q, category=category,
        btype=btype, rating=rating, sort=sort, categories=categories, types=types,
        ratings=ratings, stats=stats,
    )


@app.route("/offer/<slug>")
def offer(slug):
    b = db.get_bond(slug=slug)
    if not b:
        abort(404)
    bond = models.data_to_bond(b)
    settle_default = models.next_business_day(date.today())
    acc, acc_days = bond.accrued_interest_per100(settle_default)
    price = b["offer_price"]
    dirty = price + acc
    ytm = bond.yield_from_price(price, settle_default)
    from bondlab.analytics import current_yield, macaulay_duration, modified_duration, convexity, effective_annual_yield
    mac, _ = macaulay_duration(bond, ytm, settle_default)
    mod, _ = modified_duration(bond, ytm, settle_default)
    conv, _ = convexity(bond, ytm, settle_default)
    eff = effective_annual_yield(ytm, bond._comp)
    cfs = bond.cashflows_per100(settle_default)
    scale = bond.face_value / 100.0
    cf_rows = [
        {
            "date": d.isoformat(),
            "coupon": round(a - sum(p for rd, p in bond.redemptions if rd == d), 4),
            "redemption": round(sum(p for rd, p in bond.redemptions if rd == d), 4),
            "total": round(a, 4),
            "total_qty": round(a * scale, 2),
        }
        for d, a in cfs
    ]
    b = bond_public(b)
    try:
        cp = bond._current_period(settle_default)
        b["last_ip_date"] = cp.accrual_start.isoformat()
        b["next_ip_date"] = cp.accrual_end.isoformat()
    except Exception:
        pass
    b["accrued_days_default"] = acc_days
    b["accrued_default"] = round(acc, 6)
    b["dirty_default"] = round(dirty, 6)
    b["ytm_default"] = round(bond.xirr_from_price(price, settle_default, clean=True), 6)
    b["ytm_street_default"] = round(ytm, 6)
    b["xirr_default"] = b["ytm_default"]
    b["eff_default"] = round(eff, 6)
    b["cy_default"] = round(current_yield(bond.coupon, price), 6)
    b["mac_default"] = round(mac, 6)
    b["mod_default"] = round(mod, 6)
    b["conv_default"] = round(conv, 6)
    b["freq_num"] = bond.freq
    b["compounding"] = bond._comp
    b["day_count_used"] = bond.day_count
    b["interest_basis"] = bond.interest_basis
    b["redemptions_list"] = [[d.isoformat(), pct] for d, pct in bond.redemptions]
    return render_template("offer.html", b=b, cf_rows=cf_rows, settle_default=settle_default.isoformat(), bond_json=b)


@app.route("/b/<slug>")
def tiny(slug):
    row = db.get_bond(slug=slug)
    if not row:
        abort(404)
    if row.get("tiny_url") and row["tiny_url"].startswith("http"):
        return redirect(row["tiny_url"])
    return redirect(url_for("offer", slug=slug))


def _bond_form_json(b):
    b = dict(b)
    try:
        b["redemptions"] = json.loads(b.get("redemptions") or "[]")
    except Exception:
        b["redemptions"] = []
    b["tiny_link"] = f"/b/{b['slug']}"
    return b


@app.route("/calculator")
def calculator():
    settle_default = models.next_business_day(date.today()).isoformat()
    bonds = [_bond_form_json(b) for b in db.get_all_bonds()]
    return render_template("calculator.html", bonds=bonds, settle_default=settle_default)


def _parse_redemptions(v):
    if v is None or not str(v).strip():
        return []
    s = str(v).strip()
    try:
        val = json.loads(s)
        if isinstance(val, list):
            return [(str(d).strip(), float(p)) for d, p in val if d and p]
    except Exception:
        pass
    out = []
    for part in s.replace(";", ",").split(","):
        part = part.strip()
        if not part:
            continue
        if ":" not in part:
            continue
        d, p = part.split(":", 1)
        dp = models.parse_date(d.strip())
        if dp is None:
            continue
        out.append((dp.isoformat(), float(p)))
    return out


def _full_record(f):
    d = {
        "security_name": (f.get("security_name") or "").strip(),
        "isin": (f.get("isin") or "").strip().upper() or None,
        "coupon": models.parse_num(f.get("coupon")),
        "coupon_frequency": f.get("coupon_frequency") or "Annual",
        "face_value": models.parse_num(f.get("face_value")),
        "maturity": f.get("maturity") or None,
        "type": (f.get("type") or "").strip(),
        "issuer_category": (f.get("issuer_category") or "").strip(),
        "credit_rating": (f.get("credit_rating") or "").strip(),
        "offer_yield": models.parse_num(f.get("offer_yield")),
        "offer_price": models.parse_num(f.get("offer_price")),
        "min_investment": models.parse_min_investment(f.get("min_investment")),
        "issue_date": f.get("issue_date") or None,
        "first_coupon": f.get("first_coupon") or None,
        "call_date": f.get("call_date") or None,
        "put_date": f.get("put_date") or None,
        "day_count": f.get("day_count") or None,
        "coupon_type": (f.get("coupon_type") or "").strip(),
        "rating_agency": (f.get("rating_agency") or "").strip(),
        "guarantee": (f.get("guarantee") or "").strip(),
        "listing": (f.get("listing") or "").strip(),
        "issue_size": (f.get("issue_size") or "").strip(),
        "mode": (f.get("mode") or "").strip(),
        "sector": (f.get("sector") or "").strip(),
        "taxable": (f.get("taxable") or "").strip(),
        "series": (f.get("series") or "").strip(),
        "notes": (f.get("notes") or "").strip(),
        "redemptions": _parse_redemptions(f.get("redemptions")),
    }
    return d


def _calc_full_payload(f):
    d = _full_record(f)
    errors = []
    if not d["security_name"]:
        errors.append("Security Name required")
    if d["coupon"] is None:
        errors.append("Coupon (Interest Rate) required")
    if d["face_value"] is None or d["face_value"] <= 0:
        errors.append("Balance FV Per Bond required")
    if not d["maturity"] or models.parse_date(d["maturity"]) is None:
        errors.append("Final Maturity / Call Date required")
    if d["isin"] and len(d["isin"]) != 12:
        errors.append(f"ISIN looks invalid: {d['isin']}")
    if errors:
        raise ValueError("; ".join(errors))
    return d


def _compute_full(d, price, ytm_in, settle, qty, slab, listed, price_basis):
    from bondlab.analytics import (
        current_yield, macaulay_duration, modified_duration, convexity,
        effective_annual_yield, xirr, xnpv, yield_movement,
    )
    bond = models.data_to_bond(d)
    settle_d = models.parse_date(settle)
    if settle_d is None:
        raise ValueError("invalid settlement date")
    acc, acc_days = bond.accrued_interest_per100(settle_d)
    if price_basis == "dirty":
        clean_price = price - acc
    else:
        clean_price = price
    dirty = clean_price + acc
    if clean_price <= 0:
        raise ValueError("price must be positive")
    principal = dirty / 100.0 * bond.face_value * qty
    stamp = principal * quotesheet.STAMP_DUTY
    total = principal + stamp
    ytm = bond.yield_from_price(clean_price, settle_d)
    eff = effective_annual_yield(ytm, bond._comp)
    cur = current_yield(bond.coupon, clean_price)
    mac, _ = macaulay_duration(bond, ytm, settle_d)
    mod, _ = modified_duration(bond, ytm, settle_d)
    conv, _ = convexity(bond, ytm, settle_d)
    scale = bond.face_value / 100.0 * qty
    cfs = bond.cashflows_per100(settle_d)
    cf_rows = []
    for dte, a in cfs:
        red = sum(p for rd, p in bond.redemptions if rd == dte)
        cf_rows.append({
            "date": dte.isoformat(),
            "coupon": round(a - red, 6),
            "redemption": round(red, 6),
            "total_per_bond": round(a, 6),
            "total_qty": round(a * scale, 2),
        })
    pre_cfs = [(settle_d, -principal)] + [(dte, a * scale) for dte, a in cfs]
    xirr_pre = xirr(pre_cfs)
    xnpv_at_xirr = xnpv(xirr_pre, pre_cfs)
    xnpv_at_offer = None
    if d.get("offer_yield") is not None:
        xnpv_at_offer = xnpv(d["offer_yield"], pre_cfs)
    post_rows = []
    holding_days = (bond.maturity - settle_d).days
    cg_rate = quotesheet.LTCG_LISTED if (listed and holding_days > 365) else slab
    for dte, a in cfs:
        gross = a * scale
        coupon_part = bond.coupon / max(bond.freq, 1) * scale if bond.freq else 0.0
        net = gross - coupon_part * slab
        if dte == bond.maturity:
            red_at_mat = sum(p for rd, p in bond.redemptions if rd == bond.maturity) * scale
            gain = max(red_at_mat - total, 0.0)
            net = gross - gain * cg_rate
        post_rows.append({"date": dte.isoformat(), "gross": round(gross, 2), "net": round(net, 2)})
    post_cfs = [(settle_d, -total)] + [(models.parse_date(r["date"]), r["net"]) for r in post_rows]
    xirr_post = xirr(post_cfs)
    movement = yield_movement(bond, ytm, settle_d)
    ytc_rows = bond.yield_to_call(clean_price, settle_d) if bond.calls else []
    ytw = bond.yield_to_worst(clean_price, settle_d) if bond.calls else None
    return {
        "bond": {
            "security_name": d["security_name"],
            "isin": d["isin"],
            "coupon": bond.coupon,
            "coupon_frequency": d["coupon_frequency"],
            "freq_num": bond.freq,
            "face_value": bond.face_value,
            "maturity": bond.maturity.isoformat(),
            "day_count": bond.day_count,
            "interest_basis": bond.interest_basis,
            "compounding": bond._comp,
            "residual": models.compute_residual(bond.maturity.isoformat()),
            "redemptions_list": [[x.isoformat(), p] for x, p in bond.redemptions],
            "type": d["type"],
            "issuer_category": d["issuer_category"],
            "credit_rating": d["credit_rating"],
        },
        "settlement": settle_d.isoformat(),
        "accrued_per100": acc,
        "accrued_days": acc_days,
        "dirty_price": dirty,
        "clean_price": clean_price,
        "principal": principal,
        "stamp": stamp,
        "total": total,
        "qty": qty,
        "ytm": ytm,
        "investor_yield": xirr_pre,
        "effective_annual": eff,
        "current_yield": cur,
        "macaulay": mac,
        "modified": mod,
        "convexity": conv,
        "xirr_pre": xirr_pre,
        "xirr_post": xirr_post,
        "xnpv_at_xirr": xnpv_at_xirr,
        "xnpv_at_offer_yield": xnpv_at_offer,
        "ytc": [(dte.isoformat(), y) for dte, y in ytc_rows],
        "ytw": {"type": ytw[0], "date": ytw[1].isoformat(), "yield": ytw[2]} if ytw else None,
        "cashflows": cf_rows,
        "post_tax": post_rows,
        "movement": [{"yield": m["yield"], "price": m["price"], "delta": m["price"] - clean_price} for m in movement],
    }


@app.route("/api/calc_full", methods=["POST"])
def api_calc_full():
    d = request.get_json(force=True) or {}
    try:
        rec = _calc_full_payload(d.get("bond") or {})
        mode = d.get("mode", "price_to_yield")
        settle = d.get("settle") or models.next_business_day(date.today()).isoformat()
        qty = float(d.get("qty") or 1)
        if qty <= 0:
            raise ValueError("quantity must be positive")
        slab = float(d.get("slab", 30)) / 100.0
        listed = str(d.get("listed", "1")) in ("1", "true", "True", "yes")
        price_basis = str(d.get("price_basis", "clean"))
        yield_basis = str(d.get("yield_basis", "bey"))
        if mode == "yield_to_price":
            y = float(d.get("yield"))
            if y < -50 or y > 200:
                raise ValueError("yield out of range")
            bp = models.data_to_bond(rec)
            sd = models.parse_date(settle)
            pac, _ = bp._pricing_accrued_per100(sd)
            if yield_basis == "xirr":
                if price_basis == "dirty":
                    dirty = bp.price_from_xirr(y, sd, clean=False)
                    clean = dirty - pac
                else:
                    clean = bp.price_from_xirr(y, sd, clean=True)
            else:
                clean = bp.price_from_yield(y, sd) - pac
            out = _compute_full(rec, clean, y, settle, qty, slab, listed, "clean")
            out["input_yield"] = y
            out["yield_basis"] = yield_basis
        else:
            price = float(d.get("price"))
            if price <= 0:
                raise ValueError("price must be positive")
            out = _compute_full(rec, price, None, settle, qty, slab, listed, price_basis)
            out["input_price"] = price
            out["yield_basis"] = yield_basis
        return jsonify(out)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400


@app.route("/api/quote_full", methods=["POST"])
def api_quote_full():
    d = request.get_json(force=True) or {}
    try:
        rec = _calc_full_payload(d.get("bond") or {})
        settle = d.get("settle") or models.next_business_day(date.today()).isoformat()
        qty = float(d.get("qty") or 1)
        price = float(d.get("price"))
        slab = float(d.get("slab", 30)) / 100.0
        listed = str(d.get("listed", "1")) in ("1", "true", "True", "yes")
        price_type = str(d.get("price_type", "clean"))
        client_name = str(d.get("client_name") or "").strip()
        client_contact = str(d.get("client_contact") or "").strip()
        notes = str(d.get("notes") or "").strip()
        today = date.today()
        n = db.get_db().execute("SELECT COUNT(*) FROM quotes").fetchone()[0]
        db.get_db().close()
        quote_no = f"Q{today.strftime('%Y%m%d')}-{n + 1:04d}X"
        s = quotesheet.build_quote(
            rec, client_name, client_contact, qty, price, settle,
            slab=slab, listed=listed, quote_no=quote_no, notes=notes, price_type=price_type,
        )
        xlsx_path, pdf_path = quotesheet.generate_quote_files(s, 0)
        return jsonify({
            "quote_no": quote_no,
            "xlsx": f"/quote/calc/{quote_no}/xlsx",
            "pdf": f"/quote/calc/{quote_no}/pdf",
            "summary": {k: s[k] for k in (
                "clean_price", "accrued_per100", "accrued_days", "dirty_price", "principal",
                "stamp", "total", "ytm", "effective_annual", "xirr_pre", "xirr_post", "settlement",
            )},
        })
    except ValueError as e:
        return jsonify({"error": str(e)}), 400


@app.route("/quote/calc/<quote_no>/<fmt>")
def download_calc_quote(quote_no, fmt):
    if fmt not in ("xlsx", "pdf"):
        abort(404)
    matches = glob.glob(os.path.join(quotesheet.DATA_DIR, f"{quote_no}_*.{fmt}"))
    if not matches:
        abort(404)
    return send_file(matches[0], as_attachment=True, download_name=os.path.basename(matches[0]))


@app.route("/api/bonds")
def api_bonds():
    q = request.args.get("q", "").strip()
    category = request.args.get("category", "").strip()
    btype = request.args.get("type", "").strip()
    rating = request.args.get("rating", "").strip()
    sort = request.args.get("sort", "yield")
    bonds = db.list_bonds(q=q or None, category=category or None, btype=btype or None, rating=rating or None, sort=sort)
    return jsonify([bond_public(b) for b in bonds])


@app.route("/api/calc", methods=["POST"])
def api_calc():
    d = request.get_json(force=True)
    row = db.get_bond(slug=d.get("slug"))
    if not row:
        return jsonify({"error": "bond not found"}), 404
    try:
        settle = d.get("settle") or models.next_business_day(date.today()).isoformat()
        qty = float(d.get("qty", 1))
        price = d.get("price")
        y = d.get("yield")
        if price is not None:
            price = float(price)
        elif y is not None:
            bond = models.data_to_bond(row)
            acc, _ = bond.accrued_interest_per100(models.parse_date(settle))
            price = bond.price_from_yield(float(y), models.parse_date(settle)) - acc
        else:
            price = float(row["offer_price"])
        s = quotesheet.build_quote(
            row, d.get("client_name", ""), d.get("client_contact", ""), qty, price, settle,
            slab=float(d.get("slab", 0.30)), listed=bool(d.get("listed", True)),
            price_type=str(d.get("price_type", "clean")),
        )
        out = {k: s[k] for k in (
            "clean_price", "accrued_per100", "accrued_days", "dirty_price", "principal",
            "stamp", "total", "ytm", "effective_annual", "current_yield", "macaulay",
            "modified", "convexity", "xirr_pre", "xirr_post", "xnpv_at_xirr",
            "xnpv_at_offer_yield", "settlement", "movement", "ytc", "ytw",
        )}
        return jsonify(out)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400


@app.route("/api/quote", methods=["POST"])
def api_quote():
    d = request.get_json(force=True) or request.form
    row = db.get_bond(slug=d.get("slug"))
    if not row:
        return jsonify({"error": "bond not found"}), 404
    try:
        qty = float(d.get("qty"))
        price = float(d.get("price"))
        settle = d.get("settle") or models.next_business_day(date.today()).isoformat()
        slab = float(d.get("slab", 0.30))
        listed = str(d.get("listed", "1")) in ("1", "true", "True", "yes")
        client_name = str(d.get("client_name") or "").strip()
        client_contact = str(d.get("client_contact") or "").strip()
        notes = str(d.get("notes") or "").strip()
        today = date.today()
        n = db.get_db().execute("SELECT COUNT(*) FROM quotes").fetchone()[0]
        db.get_db().close()
        quote_no = f"Q{today.strftime('%Y%m%d')}-{n + 1:04d}"
        price_type = str(d.get("price_type", "clean"))
        s = quotesheet.build_quote(row, client_name, client_contact, qty, price, settle, slab=slab, listed=listed, quote_no=quote_no, notes=notes, price_type=price_type)
        xlsx_path, pdf_path = quotesheet.generate_quote_files(s, 0)
        db.insert_quote({
            "bond_id": row["id"], "quote_no": quote_no, "client_name": client_name,
            "client_contact": client_contact, "qty": qty, "price": price, "settle": settle,
            "principal": s["principal"], "accrued": s["accrued_per100"], "stamp": s["stamp"],
            "total": s["total"], "ytm": s["ytm"], "xirr": s["xirr_pre"],
            "xlsx_path": xlsx_path, "pdf_path": pdf_path,
        })
        return jsonify({
            "quote_no": quote_no,
            "xlsx": f"/quote/{quote_no}/download/xlsx",
            "pdf": f"/quote/{quote_no}/download/pdf",
            "summary": {k: s[k] for k in (
                "clean_price", "accrued_per100", "accrued_days", "dirty_price", "principal",
                "stamp", "total", "ytm", "effective_annual", "xirr_pre", "xirr_post", "settlement",
            )},
        })
    except ValueError as e:
        return jsonify({"error": str(e)}), 400


@app.route("/quote/<quote_no>/download/<fmt>")
def download_quote(quote_no, fmt):
    conn = db.get_db()
    row = conn.execute("SELECT * FROM quotes WHERE quote_no=?", (quote_no,)).fetchone()
    conn.close()
    if not row:
        abort(404)
    path = row["xlsx_path"] if fmt == "xlsx" else row["pdf_path"]
    if not path or not os.path.exists(path):
        abort(404)
    name = os.path.basename(path)
    return send_file(path, as_attachment=True, download_name=name)


# ---------------------------------------------------------------- admin
@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        if request.form.get("password") == ADMIN_PASSWORD:
            session["admin"] = True
            nxt = request.args.get("next") or url_for("admin_dash")
            return redirect(nxt)
        flash("Invalid password", "error")
    return render_template("admin_login.html")


@app.route("/admin/logout")
def admin_logout():
    session.pop("admin", None)
    return redirect(url_for("admin_login"))


@app.route("/admin")
@login_required
def admin_dash():
    bonds = db.get_all_bonds()
    quotes = db.list_quotes(20)
    stats = {
        "bonds": len(bonds),
        "quotes": db.get_db().execute("SELECT COUNT(*) FROM quotes").fetchone()[0],
    }
    db.get_db().close()
    return render_template("admin_dash.html", bonds=bonds, quotes=quotes, stats=stats)


def _form_to_data(f):
    data = {
        "security_name": (f.get("security_name") or "").strip(),
        "issuer_category": (f.get("issuer_category") or "").strip(),
        "isin": (f.get("isin") or "").strip().upper() or None,
        "coupon": models.parse_num(f.get("coupon")),
        "coupon_frequency": f.get("coupon_frequency") or "Annual",
        "face_value": models.parse_num(f.get("face_value")),
        "type": (f.get("type") or "").strip(),
        "credit_rating": (f.get("credit_rating") or "").strip(),
        "maturity": f.get("maturity") or None,
        "call_date": f.get("call_date") or None,
        "put_date": f.get("put_date") or None,
        "issue_date": f.get("issue_date") or None,
        "first_coupon": f.get("first_coupon") or None,
        "day_count": f.get("day_count") or None,
        "residual": (f.get("residual") or "").strip(),
        "offer_yield": models.parse_num(f.get("offer_yield")),
        "min_investment": models.parse_min_investment(f.get("min_investment")),
        "offer_price": models.parse_num(f.get("offer_price")),
        "tiny_url": (f.get("tiny_url") or "").strip(),
        "guarantee": (f.get("guarantee") or "").strip(),
        "listing": (f.get("listing") or "").strip(),
        "sector": (f.get("sector") or "").strip(),
        "taxable": (f.get("taxable") or "").strip(),
        "coupon_type": (f.get("coupon_type") or "").strip(),
        "rating_agency": (f.get("rating_agency") or "").strip(),
        "issue_size": (f.get("issue_size") or "").strip(),
        "mode": (f.get("mode") or "").strip(),
        "series": (f.get("series") or "").strip(),
        "notes": (f.get("notes") or "").strip(),
    }
    return data


@app.route("/admin/bonds/new", methods=["GET", "POST"])
@login_required
def admin_bond_new():
    if request.method == "POST":
        data = _form_to_data(request.form)
        if not data.get("offer_price"):
            models.fill_offer_price(data)
        errors = models.validate(data)
        if not errors and not data.get("offer_price") and data.get("offer_yield") is not None:
            errors.append("Offer Price could not be computed from Offer Yield; check maturity and coupon frequency")
        if not errors and data["isin"] and db.bond_exists_isin(data["isin"]):
            errors.append(f"ISIN {data['isin']} already exists")
        if errors:
            return render_template("admin_bond_form.html", b=data, errors=errors, is_new=True)
        data = models.normalize_record(data)
        data["slug"] = make_slug(data["security_name"])
        if not data["tiny_url"]:
            data["tiny_url"] = f"/b/{data['slug']}"
        db.insert_bond(data)
        flash(f"Bond '{data['security_name']}' added", "ok")
        return redirect(url_for("admin_dash"))
    return render_template("admin_bond_form.html", b={}, errors=[], is_new=True)


@app.route("/admin/bonds/<int:bid>/edit", methods=["GET", "POST"])
@login_required
def admin_bond_edit(bid):
    row = db.get_bond(bond_id=bid)
    if not row:
        abort(404)
    if request.method == "POST":
        data = _form_to_data(request.form)
        if not data.get("offer_price"):
            models.fill_offer_price(data)
        errors = models.validate(data)
        if not errors and not data.get("offer_price") and data.get("offer_yield") is not None:
            errors.append("Offer Price could not be computed from Offer Yield; check maturity and coupon frequency")
        if not errors and data["isin"] and db.bond_exists_isin(data["isin"], exclude_id=bid):
            errors.append(f"ISIN {data['isin']} already exists")
        if errors:
            return render_template("admin_bond_form.html", b=data, errors=errors, is_new=False, bid=bid)
        data = models.normalize_record(data)
        if not data["tiny_url"]:
            data["tiny_url"] = f"/b/{row['slug']}"
        db.update_bond(bid, data)
        flash(f"Bond '{data['security_name']}' updated", "ok")
        return redirect(url_for("admin_dash"))
    return render_template("admin_bond_form.html", b=row, errors=[], is_new=False, bid=bid)


@app.route("/admin/bonds/<int:bid>/delete", methods=["POST"])
@login_required
def admin_bond_delete(bid):
    db.delete_bond(bid)
    flash("Bond deleted", "ok")
    return redirect(url_for("admin_dash"))


@app.route("/admin/import", methods=["GET", "POST"])
@login_required
def admin_import():
    if request.method == "POST":
        stage = request.form.get("stage")
        if stage == "upload":
            file = request.files.get("file")
            if not file or not file.filename.lower().endswith(".xlsx"):
                flash("Please upload an .xlsx file", "error")
                return redirect(url_for("admin_import"))
            wb = openpyxl.load_workbook(io.BytesIO(file.read()), data_only=True)
            preview = []
            fmt_detected = None
            for ws in wb.worksheets:
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    continue
                header_idx = None
                for i, r in enumerate(rows):
                    vals = [str(v).strip() if v is not None else "" for v in r]
                    if any(v in models.CANONICAL_HEADERS or v in models.LEGACY_HEADERS for v in vals):
                        header_idx = i
                        break
                if header_idx is None:
                    continue
                headers = [str(v).strip() if v is not None else "" for v in rows[header_idx]]
                for ri, r in enumerate(rows[header_idx + 1:], start=header_idx + 2):
                    if all(v is None or str(v).strip() == "" for v in r):
                        continue
                    row = {headers[j]: (r[j] if j < len(r) else None) for j in range(len(headers))}
                    data, fmt = models.detect_and_map(row)
                    if fmt == "unknown":
                        continue
                    fmt_detected = fmt_detected or fmt
                    errors = models.validate(data)
                    if not data.get("offer_price") and data.get("offer_yield") is not None:
                        models.fill_offer_price(data)
                        errors = [e for e in errors if e != "Offer Price required"]
                        if not data.get("offer_price"):
                            errors.append("Offer Price could not be computed from Offer Yield; check maturity and coupon frequency")
                    if not errors and data.get("isin") and db.bond_exists_isin(data["isin"]):
                        errors.append(f"ISIN {data['isin']} already exists in database")
                    preview.append({
                        "sheet": ws.title, "row": ri, "data": data,
                        "ok": not errors, "errors": errors,
                    })
                break
            if not preview:
                flash("No bond rows found; check headers (14-column template or legacy 25-column format).", "error")
                return redirect(url_for("admin_import"))
            session["import_preview"] = preview
            return render_template("admin_import.html", preview=preview, confirm=True, fmt=fmt_detected)
        if stage == "confirm":
            preview = session.pop("import_preview", [])
            ok_rows = [p for p in preview if p["ok"]]
            inserted = 0
            for p in ok_rows:
                data = models.normalize_record(p["data"])
                data["slug"] = make_slug(data["security_name"])
                if not data["tiny_url"]:
                    data["tiny_url"] = f"/b/{data['slug']}"
                db.insert_bond(data)
                inserted += 1
            return render_template(
                "admin_import.html", preview=[], confirm=False,
                result={"inserted": inserted, "skipped": len(preview) - inserted},
            )
    return render_template("admin_import.html", preview=[], confirm=False, result=None)


@app.route("/admin/template.xlsx")
@login_required
def admin_template():
    wb = openpyxl.Workbook()
    from openpyxl.styles import Font, PatternFill
    hdr_fill = PatternFill("solid", fgColor="0B1220")
    hdr_font = Font(color="34D399", bold=True)
    ws = wb.active
    ws.title = "Bonds (14 columns)"
    for i, h in enumerate(models.CANONICAL_HEADERS, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill = hdr_fill
        c.font = hdr_font
    examples = [
        ["9.85", "Acme Fintech NCD Series III", "NBFC", "INE0ABC12345", "Annual", 1000, "NCD", "AA (Crisil)", "2032-03-20", "", 9.40, "1", 102.35, "/b/acme-fintech-ncd-series-iii"],
        ["7.18", "7.18% GS 2033", "Government", "IN0020230087", "Semi-Annual", 100, "G-Sec", "SOV", "2033-08-14", "", 6.95, "1", 101.62, "/b/718-gs-2033"],
        ["0", "ACME ZCB 2027", "Corporate", "INE0XYZ98765", "Zero Coupon", 100000, "ZCB", "A (Icra)", "2027-06-30", "", 8.80, "10", 77.50, "/b/acme-zcb-2027"],
    ]
    for ri, row in enumerate(examples, 2):
        for ci, v in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=v)
        link = ws.cell(row=ri, column=14)
        link.hyperlink = request.host_url.rstrip("/") + str(row[13])
        link.style = "Hyperlink"
    ws.column_dimensions["A"].width = 16
    for col in "BCDEFGHIJKLMN":
        ws.column_dimensions[col].width = 18

    ws2 = wb.create_sheet("Legacy (25 columns)")
    legacy_headers = list(models.LEGACY_HEADERS.keys())
    for i, h in enumerate(legacy_headers, 1):
        c = ws2.cell(row=1, column=i, value=h)
        c.fill = hdr_fill
        c.font = hdr_font
    ws2.cell(row=2, column=1, value="INE0ABC12345")
    ws2.cell(row=2, column=2, value="Acme Fintech NCD")
    ws2.cell(row=2, column=3, value="9.85%")
    ws2.cell(row=2, column=4, value="20-03-2032")
    ws2.cell(row=2, column=7, value="Annual")
    ws2.cell(row=2, column=8, value="9.40%")
    ws2.cell(row=2, column=10, value="AA")
    ws2.cell(row=2, column=12, value="NCD")
    ws2.cell(row=2, column=13, value="1000")
    ws2.cell(row=2, column=14, value="1 Lac")

    ws3 = wb.create_sheet("Instructions")
    instr = [
        "BondLab bulk import instructions",
        "",
        "Sheet 'Bonds (14 columns)' is the standard template. Headers must match exactly.",
        "Sheet 'Legacy (25 columns)' maps the original AdminWebsiteheaders format (ISIN Number, Coupon Rate, ...).",
        "",
        "Field rules:",
        "1. Coupon (Interest Rate): number in % p.a., e.g. 9.85. Use 0 for zero coupon.",
        "2. Security Name: free text, required.",
        "3. Issuer Category: e.g. NBFC, Bank, Corporate, PSU, Government.",
        "4. ISIN No.: 12 characters, e.g. INE0ABC12345. Optional but must be unique.",
        "5. Coupon Frequency: Annual | Semi-Annual | Quarterly | Monthly | Zero Coupon.",
        "6. Balance FV Per Bond: face value in rupees, e.g. 1000 or 100000.",
        "7. Type: NCD, Bond, G-Sec, SDL, SDI, ZCB, Debenture, etc.",
        "8. Credit Rating: e.g. AA (Crisil), A+ (Icra).",
        "9. Final Maturity / Call Date: dd-mm-yyyy or dd/mm/yyyy or yyyy-mm-dd or 20-Mar-2032.",
        "10. Residual Period to Maturity / Call: optional; computed from maturity if blank.",
        "11. Offer Yield Percentage: number in %, e.g. 9.40.",
        "12. Min Investment: multiples of Rs lacs; '1' = 1 lac = Rs 1,00,000. Also accepts 100000 or '2 Lacs'.",
        "13. Offer Price: clean price per Rs 100 face, e.g. 102.35.",
        "14. Tiny URL: optional; if blank, auto-generated as a link to the offer page (e.g. /b/acme-fintech-ncd-series-iii).",
        "",
        "All calculations (YTM, accrued, settlement, XIRR, duration) run server-side on the",
        "street-convention engine with machine-precision yield solving.",
    ]
    for i, line in enumerate(instr, 1):
        ws3.cell(row=i, column=1, value=line)
    ws3.column_dimensions["A"].width = 110

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(out, as_attachment=True, download_name="BondLab_bonds_template.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/admin/export.xlsx")
@login_required
def admin_export():
    bonds = db.get_all_bonds()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bonds"
    for i, h in enumerate(models.CANONICAL_HEADERS, 1):
        ws.cell(row=1, column=i, value=h)
    fmap = {
        "Coupon (Interest Rate)": "coupon",
        "Security Name": "security_name",
        "Issuer Category": "issuer_category",
        "ISIN No.": "isin",
        "Coupon Frequency": "coupon_frequency",
        "Balance FV Per Bond": "face_value",
        "Type": "type",
        "Credit Rating": "credit_rating",
        "Final Maturity / Call Date": "maturity",
        "Residual Period to Maturity / Call": "residual",
        "Offer Yield Percentage": "offer_yield",
        "Min Investment: Multiples of (₹ Lacs)": "min_investment",
        "Offer Price": "offer_price",
        "Tiny URL": "tiny_url",
    }
    for ri, b in enumerate(bonds, 2):
        for ci, h in enumerate(models.CANONICAL_HEADERS, 1):
            v = b.get(fmap[h])
            if v == "":
                v = None
            ws.cell(row=ri, column=ci, value=v)
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(out, as_attachment=True, download_name=f"BondLab_bonds_{date.today():%Y%m%d}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/admin/quotes")
@login_required
def admin_quotes():
    quotes = db.list_quotes(200)
    return render_template("admin_quotes.html", quotes=quotes)


def seed_demo():
    path = os.path.join(BASE_DIR, "examples", "real_bonds.json")
    if not os.path.exists(path):
        return
    import sqlite3
    try:
        if db.count_bonds() > 0:
            bundle_rows = json.load(open(path))
            staggered = {r["isin"]: r.get("redemptions") for r in bundle_rows if r.get("redemptions")}
            _conn = db.get_db()
            try:
                rows = _conn.execute("SELECT isin, redemptions FROM bonds").fetchall()
            finally:
                _conn.close()
            cur = {r["isin"]: json.loads(r["redemptions"] or "[]") for r in rows}
            if all(cur.get(isin) == red for isin, red in staggered.items()):
                return
            db.get_db().execute("DELETE FROM bonds")
            db.get_db().commit()
    except Exception:
        pass
    for r in json.load(open(path)):
        try:
            data = {
                "security_name": r["security_name"],
                "issuer_category": r.get("issuer_category"),
                "isin": r.get("isin"),
                "coupon": r["coupon"],
                "coupon_frequency": r["coupon_frequency"],
                "face_value": r["face_value"],
                "type": r.get("type"),
                "credit_rating": r.get("credit_rating"),
                "maturity": r["maturity"],
                "call_date": r.get("call_date"),
                "day_count": r.get("day_count"),
                "residual": models.compute_residual(r["maturity"]),
                "offer_yield": r.get("offer_yield"),
                "min_investment": r.get("min_investment"),
                "offer_price": r["offer_price"],
                "notes": r.get("source"),
                "redemptions": json.dumps(r.get("redemptions") or []),
            }
            data = models.normalize_record(data)
            data["slug"] = make_slug(data["security_name"])
            data["tiny_url"] = f"/b/{data['slug']}"
            db.insert_bond(data)
        except sqlite3.IntegrityError:
            continue
        except Exception:
            continue


seed_demo()


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 8000)), debug=False)
