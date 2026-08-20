import os
import sys
from datetime import date, datetime

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
if BASE_DIR not in sys.path:
    sys.path.insert(0, BASE_DIR)
from bondlab import Bond
from bondlab.analytics import (
    current_yield,
    macaulay_duration,
    modified_duration,
    convexity,
    yield_movement,
    xirr,
    xnpv,
    effective_annual_yield,
)

from models import data_to_bond, parse_date

STAMP_DUTY = 0.000001
LTCG_LISTED = 0.125

DATA_DIR = os.environ.get("BONDAPP_GEN_DIR") or (
    "/tmp/bondapp_generated" if os.environ.get("VERCEL") == "1"
    else os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "generated"))
os.makedirs(DATA_DIR, exist_ok=True)

_LOGO = os.path.join(os.path.dirname(os.path.abspath(__file__)), "static", "logo_flattened.png")

INK = "1B1813"
GOLD = "B08D2E"
GOLD_SOFT = "F6EFDD"
GOLD_LINE = "E3D6B3"
CREAM = "FCFBF7"
MUTED = "8A857A"
WHITE = "FFFFFF"
GREEN = "1A7F37"
RED = "B3261E"


def build_quote(record, client_name, client_contact, qty, price, settle, slab=0.30, listed=True, quote_no=None, notes="", price_type="clean"):
    bond = data_to_bond(record)
    settle_d = parse_date(settle)
    if settle_d is None:
        raise ValueError("invalid settlement date")
    if price_type not in ("clean", "dirty"):
        raise ValueError("price_type must be 'clean' or 'dirty'")

    accrued_per100, accrued_days = bond.accrued_interest_per100(settle_d)
    if price_type == "dirty":
        clean_price = price - accrued_per100
    else:
        clean_price = price
    dirty = clean_price + accrued_per100
    principal = dirty / 100.0 * bond.face_value * qty
    stamp = principal * STAMP_DUTY
    total = principal + stamp
    ytm = bond.yield_from_price(clean_price, settle_d)
    eff = effective_annual_yield(ytm, bond._comp)
    cur = current_yield(bond.coupon, price)
    ytc_rows = bond.yield_to_call(clean_price, settle_d) if bond.calls else []
    ytw = bond.yield_to_worst(clean_price, settle_d) if bond.calls else None
    mac, _ = macaulay_duration(bond, ytm, settle_d)
    mod, _ = modified_duration(bond, ytm, settle_d)
    conv, _ = convexity(bond, ytm, settle_d)

    scale = bond.face_value / 100.0 * qty
    cfs = bond.cashflows_per100(settle_d)

    def row_for(d, a):
        red = sum(pct for rd, pct in bond.redemptions if rd == d)
        coupon_amt = a - red
        return {
            "date": d.isoformat(),
            "coupon": coupon_amt,
            "redemption": red,
            "total_per_bond": a,
            "total_qty": a * scale,
        }

    cf_rows = [row_for(d, a) for d, a in cfs]

    pre_cfs = [(settle_d, -total)] + [(d, a * scale) for d, a in cfs]
    xirr_pre = xirr(pre_cfs)
    xnpv_at_xirr = xnpv(xirr_pre, pre_cfs)
    offer_y = record.get("offer_yield")
    xnpv_at_offer = xnpv(offer_y, pre_cfs) if offer_y is not None else None

    post_rows = []
    holding_days = (bond.maturity - settle_d).days
    cg_rate = LTCG_LISTED if (listed and holding_days > 365) else slab
    for d, a in cfs:
        gross = a * scale
        coupon_part = bond.coupon / max(bond.freq, 1) * scale if bond.freq else 0.0
        net = gross - coupon_part * slab
        if d == bond.maturity:
            red_at_mat = sum(pct for rd, pct in bond.redemptions if rd == bond.maturity) * scale
            gain = max(red_at_mat - total, 0.0)
            net = gross - gain * cg_rate
        post_rows.append({"date": d.isoformat(), "gross": gross, "net": net})
    post_cfs = [(settle_d, -total)] + [(parse_date(r["date"]), r["net"]) for r in post_rows]
    xirr_post = xirr(post_cfs)

    movement = yield_movement(bond, ytm, settle_d)

    summary = {
        "quote_no": quote_no,
        "generated": date.today().isoformat(),
        "client_name": client_name,
        "client_contact": client_contact,
        "notes": notes,
        "price_type": price_type,
        "bond": record,
        "coupon": bond.coupon,
        "freq": bond.freq,
        "day_count": bond.day_count,
        "compounding": bond._comp,
        "settlement": settle_d.isoformat(),
        "ytc": [(d.isoformat(), y) for d, y in ytc_rows],
        "ytw": {"type": ytw[0], "date": ytw[1].isoformat(), "yield": ytw[2]} if ytw else None,
        "clean_price": clean_price,
        "accrued_per100": accrued_per100,
        "accrued_days": accrued_days,
        "dirty_price": dirty,
        "qty": qty,
        "principal": principal,
        "stamp": stamp,
        "total": total,
        "ytm": ytm,
        "effective_annual": eff,
        "current_yield": cur,
        "macaulay": mac,
        "modified": mod,
        "convexity": conv,
        "xirr_pre": xirr_pre,
        "xnpv_at_xirr": xnpv_at_xirr,
        "xnpv_at_offer_yield": xnpv_at_offer,
        "xirr_post": xirr_post,
        "slab": slab,
        "listed": listed,
        "cg_rate": cg_rate,
        "cashflows": cf_rows,
        "post_tax": post_rows,
        "movement": movement,
        "maturity": bond.maturity.isoformat(),
        "residual_days": (bond.maturity - settle_d).days,
    }
    return summary


def _sanitize(name):
    return "".join(ch for ch in name if ch.isalnum() or ch in "._- ")[:60].strip() or "bond"


def generate_excel(s, path):
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.worksheet.properties import PageSetupProperties

    b = s["bond"]
    wb = openpyxl.Workbook()

    fill_ink = PatternFill("solid", fgColor=INK)
    fill_gold_soft = PatternFill("solid", fgColor=GOLD_SOFT)
    fill_cream = PatternFill("solid", fgColor=CREAM)
    thin = Side(style="thin", color=GOLD_LINE)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    left_gold = Border(left=Side(style="thick", color=GOLD), right=thin, top=thin, bottom=thin)
    money = '₹#,##0.00'
    price4 = '0.0000'
    yield4 = '0.0000"%"'
    delta4 = '+0.0000;-0.0000;0.0000'

    def page_setup(ws):
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)

    def masthead(ws, ncols):
        last = openpyxl.utils.get_column_letter(ncols)
        for r in range(1, 4):
            ws.row_dimensions[r].height = 30
        ws.merge_cells(f"A1:{last}3")
        c = ws.cell(row=1, column=1)
        c.value = "BONDLAB | BOND QUOTE SHEET"
        c.fill = fill_ink
        c.font = Font(color=WHITE, bold=True, size=20)
        c.alignment = Alignment(horizontal="center", vertical="center")
        if os.path.exists(_LOGO):
            from openpyxl.drawing.image import Image as XLImage
            img = XLImage(_LOGO)
            img.width = 58
            img.height = 58
            ws.add_image(img, "A1")
        ws.merge_cells(f"A4:{last}4")
        ws.row_dimensions[4].height = 22
        m = ws.cell(row=4, column=1)
        m.value = (
            f"Quote No: {s['quote_no']}    •    Generated: {s['generated']}    •    "
            f"Settlement: {s['settlement']}    •    Client: {s['client_name'] or ''}"
        )
        m.font = Font(color=GOLD, size=10)
        m.alignment = Alignment(horizontal="center", vertical="center")

    def section(ws, r, label, span=2):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=span)
        c = ws.cell(row=r, column=1)
        c.value = label
        c.fill = fill_gold_soft
        c.font = Font(color=INK, bold=True, size=10.5)
        c.border = left_gold
        c.alignment = Alignment(vertical="center")
        ws.row_dimensions[r].height = 22
        return r + 1

    def kv(ws, r, k, v, vfmt=None, total=False, italic=False):
        c1 = ws.cell(row=r, column=1, value=k)
        c2 = ws.cell(row=r, column=2, value=v)
        if total:
            c1.fill = fill_ink
            c2.fill = fill_ink
            c1.font = Font(color=GOLD, bold=True, size=10.5)
            c2.font = Font(color=GOLD, bold=True, size=10.5)
        else:
            if r % 2 == 1:
                c1.fill = fill_cream
                c2.fill = fill_cream
            c1.font = Font(color=MUTED, size=10, italic=italic)
            c2.font = Font(color=INK, bold=True, size=10)
        c1.border = border
        c2.border = border
        c2.alignment = Alignment(horizontal="right")
        if vfmt:
            c2.number_format = vfmt
        ws.row_dimensions[r].height = 17
        return r + 1

    def data_title(ws, title, ncols):
        last = openpyxl.utils.get_column_letter(ncols)
        ws.merge_cells(f"A1:{last}1")
        t = ws.cell(row=1, column=1, value=f"{title}  |  {s['quote_no']}")
        t.fill = fill_ink
        t.font = Font(color=GOLD, bold=True, size=12)
        t.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

    def data_header(ws, heads):
        for i, h in enumerate(heads, 1):
            c = ws.cell(row=2, column=i, value=h)
            c.fill = fill_ink
            c.font = Font(color=WHITE, bold=True, size=10)
            c.border = border
            c.alignment = Alignment(horizontal="center" if i == 1 else "right")
        ws.row_dimensions[2].height = 18

    def band_row(ws, r, ncols, label=None, value=None, vfmt=None):
        for i in range(1, ncols + 1):
            c = ws.cell(row=r, column=i)
            c.fill = fill_gold_soft
            c.border = Border(top=Side(style="thick", color=GOLD), bottom=thin)
        if label is not None:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols - 2)
            lc = ws.cell(row=r, column=1, value=label)
            lc.font = Font(bold=True, size=10, color=INK)
            lc.alignment = Alignment(horizontal="right", vertical="center")
        if value is not None:
            vc = ws.cell(row=r, column=ncols, value=value)
            vc.font = Font(bold=True, color=GOLD, size=10.5)
            vc.alignment = Alignment(horizontal="right", vertical="center")
            if vfmt:
                vc.number_format = vfmt
        ws.row_dimensions[r].height = 18
        return r + 1

    # ---------------- Quote Summary ----------------
    ws = wb.active
    ws.title = "Quote Summary"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 26
    masthead(ws, 2)

    r = 6
    r = section(ws, r, "CLIENT")
    r = kv(ws, r, "Client Name", s["client_name"] or "")
    r = kv(ws, r, "Contact", s["client_contact"] or "")
    r += 1

    r = section(ws, r, "SECURITY")
    r = kv(ws, r, "Security Name", b["security_name"])
    r = kv(ws, r, "ISIN", b["isin"] or "")
    r = kv(ws, r, "Issuer Category", b["issuer_category"] or "")
    r = kv(ws, r, "Type", b["type"] or "")
    r = kv(ws, r, "Credit Rating", b["credit_rating"] or "")
    r = kv(ws, r, "Coupon Rate (% p.a.)", round(b["coupon"], 4), yield4)
    r = kv(ws, r, "Coupon Frequency", b["coupon_frequency"])
    r = kv(ws, r, "Face Value per Bond", b["face_value"], money)
    r = kv(ws, r, "Final Maturity / Call", s["maturity"])
    r = kv(ws, r, "Residual (days from settlement)", s["residual_days"], '#,##0')
    r = kv(ws, r, "Day Count Convention", s["day_count"])
    r += 1

    r = section(ws, r, "PRICING & SETTLEMENT BREAKUP")
    r = kv(ws, r, "Quantity (bonds)", s["qty"], '#,##0')
    r = kv(ws, r, "Settlement Date", s["settlement"])
    r = kv(ws, r, "Clean Price (per ₹100)", round(s["clean_price"], 6), price4)
    r = kv(ws, r, f"Accrued Interest (per ₹100, {s['accrued_days']} days)", round(s["accrued_per100"], 6), price4)
    r = kv(ws, r, "Dirty Price (per ₹100)", round(s["dirty_price"], 6), price4)
    r = kv(ws, r, "Principal Consideration", round(s["principal"], 2), money)
    r = kv(ws, r, "Stamp Duty (0.0001%)", round(s["stamp"], 2), money)
    r = kv(ws, r, "TOTAL SETTLEMENT AMOUNT", round(s["total"], 2), money, total=True)
    r += 1

    r = section(ws, r, "RISK & RETURN METRICS")
    r = kv(ws, r, "Yield to Maturity (YTM, nominal)", round(s["ytm"], 4), yield4)
    for cd, yc in s["ytc"]:
        r = kv(ws, r, f"Yield to Call ({cd})", round(yc, 4), yield4)
    if s["ytw"]:
        r = kv(ws, r, f"Yield to Worst ({s['ytw']['date']})", round(s["ytw"]["yield"], 4), yield4)
    r = kv(ws, r, "Effective Annual Yield", round(s["effective_annual"], 4), yield4)
    r = kv(ws, r, "Current Yield", round(s["current_yield"], 4), yield4)
    r = kv(ws, r, "Macaulay Duration (years)", round(s["macaulay"], 4), price4)
    r = kv(ws, r, "Modified Duration", round(s["modified"], 4), price4)
    r = kv(ws, r, "Convexity", round(s["convexity"], 4), price4)
    r = kv(ws, r, "XIRR (pre-tax)", round(s["xirr_pre"], 4), yield4)
    r = kv(ws, r, "XNPV @ XIRR (check ≈ 0)", round(s["xnpv_at_xirr"], 2), money)
    if s["xnpv_at_offer_yield"] is not None:
        r = kv(ws, r, "XNPV @ platform offer yield", round(s["xnpv_at_offer_yield"], 2), money)
    r = kv(ws, r, f"XIRR (post-tax, slab {s['slab']*100:.0f}%)", round(s["xirr_post"], 4), yield4)
    if s["notes"]:
        r += 1
        r = kv(ws, r, "Notes", s["notes"], italic=True)
    r += 1

    ws.merge_cells(start_row=r, start_column=1, end_row=r + 3, end_column=2)
    d = ws.cell(row=r, column=1, value=(
        "Disclaimer: This quote sheet is an analytical output of the BondLab engine (street-convention bond mathematics). "
        "It is not investment advice or a solicitation. Fixed returns are not guaranteed; corporate bonds carry credit, "
        "market and liquidity risk. Verify cashflows against the ISIN master and offer documents before investing. "
        "Tax treatment is indicative (India, post Jul-2024 listed LTCG 12.5% without indexation); consult your tax advisor."
    ))
    d.font = Font(color=MUTED, size=9)
    d.alignment = Alignment(wrap_text=True, vertical="top")
    page_setup(ws)

    # ---------------- Cashflows ----------------
    ws2 = wb.create_sheet("Cashflows")
    ws2.sheet_view.showGridLines = False
    heads = ["Date", "Coupon (per bond)", "Redemption (per bond)", "Total (per bond)", "Total (all bonds)"]
    data_title(ws2, "CASHFLOW SCHEDULE", len(heads))
    data_header(ws2, heads)
    for ri, row in enumerate(s["cashflows"]):
        rr = ri + 3
        vals = [row["date"], round(row["coupon"], 4), round(row["redemption"], 4), round(row["total_per_bond"], 4), round(row["total_qty"], 2)]
        for i, v in enumerate(vals, 1):
            c = ws2.cell(row=rr, column=i, value=v)
            c.border = border
            if rr % 2 == 1:
                c.fill = fill_cream
            if i > 1:
                c.number_format = money if i == 5 else price4
                c.alignment = Alignment(horizontal="right")
    r = len(s["cashflows"]) + 4
    r = band_row(ws2, r, 5, label="XIRR (pre-tax) of settlement outflow + cashflows",
                 value=round(s["xirr_pre"], 4), vfmt=yield4)
    for col, w in zip("ABCDE", [15, 21, 21, 19, 21]):
        ws2.column_dimensions[col].width = w
    ws2.freeze_panes = "A3"
    page_setup(ws2)

    # ---------------- Post-Tax ----------------
    ws3 = wb.create_sheet("Post-Tax")
    ws3.sheet_view.showGridLines = False
    heads3 = ["Date", "Pre-Tax Cashflow", "Post-Tax Cashflow", "Tax Applied"]
    data_title(ws3, "POST-TAX CASHFLOWS", len(heads3))
    data_header(ws3, heads3)
    rr = 3
    vals = [f"Outflow (settlement, {s['settlement']})", -round(s["total"], 2), -round(s["total"], 2), ""]
    for i, v in enumerate(vals, 1):
        c = ws3.cell(row=rr, column=i, value=v)
        c.border = border
        if rr % 2 == 1:
            c.fill = fill_cream
        if i in (2, 3):
            c.number_format = money
            c.alignment = Alignment(horizontal="right")
    for ri, row in enumerate(s["post_tax"]):
        rr = ri + 4
        tax_applied = f"coupon @ slab {s['slab']*100:.0f}%"
        if row["date"] == s["maturity"]:
            tax_applied = f"coupon @ slab + gain @ {s['cg_rate']*100:.2f}%"
        vals = [row["date"], round(row["gross"], 2), round(row["net"], 2), tax_applied]
        for i, v in enumerate(vals, 1):
            c = ws3.cell(row=rr, column=i, value=v)
            c.border = border
            if rr % 2 == 1:
                c.fill = fill_cream
            if i in (2, 3):
                c.number_format = money
                c.alignment = Alignment(horizontal="right")
    r = len(s["post_tax"]) + 5
    band_row(ws3, r, 4, label=f"XIRR (post-tax, slab {s['slab']*100:.0f}%)",
             value=round(s["xirr_post"], 4), vfmt=yield4)
    for col, w in zip("ABCD", [16, 20, 20, 36]):
        ws3.column_dimensions[col].width = w
    ws3.freeze_panes = "A3"
    page_setup(ws3)

    # ---------------- Yield Sensitivity ----------------
    ws4 = wb.create_sheet("Yield Sensitivity")
    ws4.sheet_view.showGridLines = False
    heads4 = ["Yield (%)", "Clean Price (per ₹100)", "Δ Price", f"Settlement Total @ qty {s['qty']:.0f}"]
    data_title(ws4, "YIELD SENSITIVITY", len(heads4))
    data_header(ws4, heads4)
    base = round(s["ytm"], 4)
    for ri, mv in enumerate(s["movement"]):
        rr = ri + 3
        cp = mv["price"]
        tot_mv = (cp + s["accrued_per100"]) / 100.0 * s["bond"]["face_value"] * s["qty"] * (1 + STAMP_DUTY)
        vals = [mv["yield"], cp, cp - s["clean_price"], round(tot_mv, 2)]
        quote_row = abs(mv["yield"] - base) < 1e-9
        for i, v in enumerate(vals, 1):
            c = ws4.cell(row=rr, column=i, value=v)
            c.border = border
            if quote_row:
                c.fill = fill_gold_soft
            elif rr % 2 == 1:
                c.fill = fill_cream
            if i == 1:
                c.number_format = yield4
            if i in (2, 3):
                c.number_format = delta4 if i == 3 else price4
                if i == 3:
                    c.font = Font(color=GREEN if v > 0 else (RED if v < 0 else INK), size=10)
            if i == 4:
                c.number_format = money
            if i > 1:
                c.alignment = Alignment(horizontal="right")
    r = len(s["movement"]) + 4
    c = ws4.cell(row=r, column=1, value="Highlighted row = quoted yield")
    c.font = Font(color=MUTED, size=9, italic=True)
    for col, w in zip("ABCD", [14, 24, 14, 30]):
        ws4.column_dimensions[col].width = w
    ws4.freeze_panes = "A3"
    page_setup(ws4)

    wb.save(path)
    return path


def generate_pdf(s, path):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, LongTable,
        Image, HRFlowable,
    )
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.pdfgen import canvas as _canvas

    _BUNDLED = os.path.join(BASE_DIR, "static", "fonts")
    _DV = "/usr/share/fonts/truetype/dejavu"
    dv = os.path.join(_BUNDLED, "DejaVuSans.ttf") if os.path.exists(os.path.join(_BUNDLED, "DejaVuSans.ttf")) else os.path.join(_DV, "DejaVuSans.ttf")
    dvb = os.path.join(_BUNDLED, "DejaVuSans-Bold.ttf") if os.path.exists(os.path.join(_BUNDLED, "DejaVuSans-Bold.ttf")) else os.path.join(_DV, "DejaVuSans-Bold.ttf")
    if os.path.exists(dv) and os.path.exists(dvb):
        pdfmetrics.registerFont(TTFont("DejaVuSans", dv))
        pdfmetrics.registerFont(TTFont("DejaVuSans-Bold", dvb))
        _RS = '<font name="DejaVuSans">₹</font>'
    else:
        _RS = "Rs."
    _RSL = _RS  # reuse in plain label text

    b = s["bond"]
    INK = colors.HexColor("#1B1813")
    GOLD = colors.HexColor("#B08D2E")
    GOLD_LINE = colors.HexColor("#E3D6B3")
    CREAM = colors.HexColor("#FCFBF7")
    MUTED = colors.HexColor("#8A857A")
    WHITE = colors.white

    st_brand = ParagraphStyle("brand", fontName="Helvetica-Bold", fontSize=18, leading=21, textColor=WHITE)
    st_brand_sub = ParagraphStyle("brandsub", fontName="Helvetica-Bold", fontSize=9, leading=11, textColor=GOLD)
    st_meta = ParagraphStyle("meta", fontName="Helvetica", fontSize=8, leading=10.5, textColor=WHITE, alignment=2)
    st_sec = ParagraphStyle("sec", fontName="Helvetica-Bold", fontSize=10.5, leading=13, textColor=INK)
    st_cell = ParagraphStyle("cell", fontName="Helvetica", fontSize=8.5, leading=11, textColor=INK)
    st_cell_b = ParagraphStyle("cellb", fontName="Helvetica-Bold", fontSize=8.5, leading=11, textColor=INK)
    st_cell_l = ParagraphStyle("celll", fontName="Helvetica", fontSize=8.5, leading=11, textColor=MUTED)
    st_cell_gold = ParagraphStyle("cellgold", fontName="Helvetica-Bold", fontSize=9, leading=11, textColor=GOLD)
    st_disc = ParagraphStyle("disc", fontName="Helvetica", fontSize=7.5, leading=9.5, textColor=MUTED)

    class NumberedCanvas(_canvas.Canvas):
        def __init__(self, *args, **kwargs):
            _canvas.Canvas.__init__(self, *args, **kwargs)
            self._saved = []

        def showPage(self):
            self._saved.append(dict(self.__dict__))
            self._startPage()

        def save(self):
            total = len(self._saved)
            for state in self._saved:
                self.__dict__.update(state)
                self._footer(total)
                _canvas.Canvas.showPage(self)
            _canvas.Canvas.save(self)

        def _footer(self, total):
            self.saveState()
            self.setStrokeColor(GOLD_LINE)
            self.setLineWidth(0.6)
            self.line(14 * mm, 12 * mm, 196 * mm, 12 * mm)
            self.setFont("Helvetica", 7.5)
            self.setFillColor(MUTED)
            self.drawString(14 * mm, 8.5 * mm, "BONDLAB  |  BOND QUOTE SHEET")
            self.drawRightString(196 * mm, 8.5 * mm, f"Page {self._pageNumber} of {total}")
            self.restoreState()

    doc = SimpleDocTemplate(path, pagesize=A4, leftMargin=14 * mm, rightMargin=14 * mm,
                            topMargin=12 * mm, bottomMargin=16 * mm)
    story = []

    def money(x, dec=2):
        return f"{_RS} {x:,.{dec}f}"

    def hdr_table():
        img = Image(_LOGO, width=13 * mm, height=13 * mm) if os.path.exists(_LOGO) else Spacer(1, 13 * mm)
        title_cell = [Paragraph("BONDLAB", st_brand), Paragraph("BOND QUOTE SHEET", st_brand_sub)]
        meta_lines = (
            f"Quote No: {s['quote_no']}<br/>"
            f"Generated: {s['generated']}<br/>"
            f"Settlement: {s['settlement']}"
        )
        t = Table([[img, title_cell, Paragraph(meta_lines, st_meta)]], colWidths=[18 * mm, 82 * mm, 80 * mm])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), INK),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 9),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 9),
            ("LEFTPADDING", (2, 0), (2, 0), 10),
            ("LINEBELOW", (0, 0), (-1, 0), 2.5, GOLD),
        ]))
        return t

    def section(label):
        t = Table([[Paragraph(label, st_sec)]], colWidths=[180 * mm])
        t.setStyle(TableStyle([
            ("LINEBEFORE", (0, 0), (0, 0), 2.5, GOLD),
            ("LEFTPADDING", (0, 0), (0, 0), 6),
            ("TOPPADDING", (0, 0), (0, 0), 10),
            ("BOTTOMPADDING", (0, 0), (0, 0), 3),
        ]))
        return t

    def kv_table(rows, total_idx=None):
        data = []
        for i, (k, v) in enumerate(rows):
            data.append([Paragraph(k, st_cell_l), Paragraph(str(v), st_cell_gold if i == total_idx else st_cell_b)])
        t = Table(data, colWidths=[88 * mm, 92 * mm])
        cmds = [
            ("GRID", (0, 0), (-1, -1), 0.4, GOLD_LINE),
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ("TOPPADDING", (0, 0), (-1, -1), 3.5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3.5),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ]
        for i in range(len(rows)):
            if i == total_idx:
                cmds.append(("BACKGROUND", (0, i), (-1, i), INK))
            elif i % 2 == 1:
                cmds.append(("BACKGROUND", (0, i), (-1, i), CREAM))
        t.setStyle(TableStyle(cmds))
        return t

    def grid_table(heads, rows, widths, right_cols=(), highlight_idx=None):
        data = [[Paragraph(h, st_cell_b) for h in heads]]
        for row in rows:
            data.append([Paragraph(str(v), st_cell) for v in row])
        t = LongTable(data, colWidths=widths, repeatRows=1)
        cmds = [
            ("GRID", (0, 0), (-1, -1), 0.4, GOLD_LINE),
            ("BACKGROUND", (0, 0), (-1, 0), INK),
            ("TEXTCOLOR", (0, 0), (-1, 0), GOLD),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ]
        for c in right_cols:
            cmds.append(("ALIGN", (c, 0), (c, -1), "RIGHT"))
        for i in range(1, len(data)):
            if i == highlight_idx:
                cmds.append(("BACKGROUND", (0, i), (-1, i), GOLD_LINE))
            elif i % 2 == 0:
                cmds.append(("BACKGROUND", (0, i), (-1, i), CREAM))
        t.setStyle(TableStyle(cmds))
        return t

    story.append(hdr_table())
    story.append(Spacer(1, 8))

    story.append(section("CLIENT"))
    story.append(kv_table([
        ("Client Name", s["client_name"] or ""),
        ("Contact", s["client_contact"] or ""),
    ]))

    story.append(section("SECURITY"))
    story.append(kv_table([
        ("Security Name", b["security_name"]),
        ("ISIN", b["isin"] or ""),
        ("Issuer Category", b["issuer_category"] or ""),
        ("Type / Credit Rating", f"{b['type'] or ''} / {b['credit_rating'] or ''}"),
        ("Coupon (% p.a.) / Frequency", f"{b['coupon']:.4f}% / {b['coupon_frequency']}"),
        ("Face Value per Bond", money(b["face_value"])),
        ("Final Maturity / Call", s["maturity"]),
        ("Residual (days)", s["residual_days"]),
        ("Day Count Convention", s["day_count"]),
    ]))

    story.append(section("PRICING & SETTLEMENT BREAKUP"))
    story.append(kv_table([
        ("Quantity (bonds)", f"{s['qty']:,.0f}"),
        ("Settlement Date", s["settlement"]),
        (f"Clean Price (per {_RSL}100)", f"{s['clean_price']:.4f}"),
        (f"Accrued Interest ({s['accrued_days']} days, per {_RSL}100)", f"{s['accrued_per100']:.4f}"),
        (f"Dirty Price (per {_RSL}100)", f"{s['dirty_price']:.4f}"),
        ("Principal Consideration", money(s["principal"])),
        ("Stamp Duty (0.0001%)", money(s["stamp"])),
        ("TOTAL SETTLEMENT AMOUNT", money(s["total"])),
    ], total_idx=7))

    story.append(section("RISK & RETURN METRICS"))
    metric_rows = [
        ("YTM (nominal)", f"{s['ytm']:.4f}%"),
        ("Effective Annual Yield", f"{s['effective_annual']:.4f}%"),
        ("Current Yield", f"{s['current_yield']:.4f}%"),
        ("Macaulay Duration (yrs)", f"{s['macaulay']:.4f}"),
        ("Modified Duration", f"{s['modified']:.4f}"),
        ("Convexity", f"{s['convexity']:.4f}"),
        ("XIRR (pre-tax)", f"{s['xirr_pre']:.4f}%"),
        ("XNPV @ XIRR (check ≈ 0)", money(s["xnpv_at_xirr"])),
        ("XNPV @ platform offer yield", money(s["xnpv_at_offer_yield"]) if s["xnpv_at_offer_yield"] is not None else ""),
        (f"XIRR (post-tax, slab {s['slab']*100:.0f}%)", f"{s['xirr_post']:.4f}%"),
    ]
    for cd, yc in s["ytc"]:
        metric_rows.insert(1, (f"Yield to Call ({cd})", f"{yc:.4f}%"))
    if s["ytw"]:
        metric_rows.insert(1, (f"Yield to Worst ({s['ytw']['date']})", f"{s['ytw']['yield']:.4f}%"))
    story.append(kv_table(metric_rows))
    if s["notes"]:
        story.append(kv_table([("Notes", s["notes"])]))

    story.append(section("CASHFLOW SCHEDULE"))
    rows = []
    for r in s["cashflows"]:
        rows.append([r["date"], f"{r['coupon']:.4f}", f"{r['redemption']:.4f}", f"{r['total_per_bond']:.4f}", f"{r['total_qty']:,.2f}"])
    story.append(grid_table(
        ["Date", "Coupon/bond", "Redemption/bond", "Total/bond", "Total (qty)"],
        rows, [30 * mm, 26 * mm, 30 * mm, 26 * mm, 30 * mm], right_cols=(1, 2, 3, 4),
    ))
    story.append(Spacer(1, 3))
    story.append(Paragraph(f"XIRR (pre-tax) of settlement outflow + cashflows: <b>{s['xirr_pre']:.4f}%</b>", st_cell))

    story.append(section("YIELD SENSITIVITY"))
    base = round(s["ytm"], 4)
    rows = []
    hi = None
    for i, m in enumerate(s["movement"]):
        rows.append([f"{m['yield']:.4f}%", f"{m['price']:.4f}", f"{m['price'] - s['clean_price']:+.4f}"])
        if abs(m["yield"] - base) < 1e-9:
            hi = i + 1
    story.append(grid_table(["Yield", "Clean Price", "Δ vs quote"], rows, [45 * mm, 45 * mm, 45 * mm],
                            right_cols=(1, 2), highlight_idx=hi))

    story.append(Spacer(1, 10))
    story.append(HRFlowable(width="100%", thickness=1, color=GOLD_LINE, spaceBefore=2, spaceAfter=6))
    story.append(Paragraph(
        "<b>Disclaimer:</b> This quote sheet is an analytical output of the BondLab engine (street-convention bond "
        "mathematics, hybrid Newton-bisection yield solving). It is not investment advice or a solicitation. Fixed "
        "returns are not guaranteed; corporate bonds carry credit, market and liquidity risk. Verify cashflows against "
        "the ISIN master and offer documents before investing. Tax treatment is indicative (India, post Jul-2024 listed "
        "LTCG 12.5% without indexation); consult your tax advisor.",
        st_disc,
    ))

    doc.build(story, canvasmaker=NumberedCanvas)
    return path


def generate_quote_files(s, quote_id):
    b = s["bond"]
    stem = f"{s['quote_no']}_{_sanitize(b['security_name'])}"
    xlsx_path = os.path.join(DATA_DIR, stem + ".xlsx")
    pdf_path = os.path.join(DATA_DIR, stem + ".pdf")
    generate_excel(s, xlsx_path)
    generate_pdf(s, pdf_path)
    return xlsx_path, pdf_path